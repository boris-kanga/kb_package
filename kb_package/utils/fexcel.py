from __future__ import annotations
import os.path
import pathlib
import re

from openpyxl.cell import Cell
import pandas
from copy import copy

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.utils import column_index_from_string, get_column_letter

from kb_package.utils.fdataset import DatasetFactory
from kb_package.tools import get_no_filepath, Cdict, replace_quoted_text, CTemporaryFile


class ExcelFactory:
    MAX_COL = "XFD"

    def __init__(self, arg=get_no_filepath("excel_temp.xlsx"), **kwargs):
        self._path = get_no_filepath("excel_temp.xlsx")
        self._kwargs = kwargs
        self._got_modification = True

        if hasattr(arg, "dataset"):
            # DatasetFactory object
            arg = arg.dataset
        if isinstance(arg, (str, pathlib.Path)):
            self._path = str(arg)
            if os.path.exists(self._path):
                if self._path.lower().endswith(".xlsm"):
                    kwargs["keep_vba"] = True
                wb = self.__load(path=arg, **kwargs)
            else:
                wb = Workbook()
        elif isinstance(arg, pandas.DataFrame):
            wb = Workbook()
            sh = wb.active
            for row in dataframe_to_rows(arg, index=False, header=True):
                sh.append(row)
        else:
            raise ValueError("Got bad value: %s" % (arg,))
        self._workbook: Workbook = wb
        self._tables_ref = []
        self._pivots_ref = []
        self._load_context()

        self._cache_copy = None
        self._cm_force_saving = True

    def __load(self, path=None, **kwargs):
        _kwargs = self._kwargs.copy()
        _kwargs.update(kwargs)

        path = self._path if not path else path
        self._workbook = load_workbook(path, **{k: v for k, v in _kwargs.items()
                                                if k in load_workbook.__code__.co_varnames})
        self._got_modification = False
        return self._workbook

    def pivot(self, name, get_idx=True, **kwargs):
        for pivot in self._pivots_ref:
            if name == pivot.name:
                if get_idx:
                    return pivot, self._workbook[pivot.sheet]._pivots[pivot.index]
                return pivot, self._workbook[pivot.sheet]._pivots[pivot.index]
        if "default" in kwargs:
            return kwargs["default"]
        raise ValueError("pivot named %s not exists" % name)

    def _get_ref_sheet_from(self, ref, sheet=None):
        if ref in self.tables_name:
            idx, table = self.table(ref, get_idx=True, default=-1)
            ref = table.ref
            sheet = self._tables_ref[idx].sheet
        ref = self._parse_ref(ref)
        if sheet is None:
            sheet = ref.get("sheet")
        if sheet is None:
            sheet = self._workbook.active
        elif not isinstance(sheet, Worksheet):
            sheet = self.sheet(sheet)
        return ref, sheet

    def ref_to_dataframe(self, ref, sheet=None):
        ref, sheet = self._get_ref_sheet_from(ref, sheet)

        columns = list(sheet.iter_rows(
            min_row=ref.start.line, max_row=ref.start.line,
            min_col=ref.start.col_id, max_col=ref.end.col_id,
            values_only=True))[0]
        dataset = pandas.DataFrame([
            list(d) for d in sheet.iter_rows(min_row=ref.start.line + 1, max_row=ref.end.line,
                                             min_col=ref.start.col_id, max_col=ref.end.col_id, values_only=True)
        ], columns=columns)

        return dataset

    @property
    def tables_name(self):
        return [d.name for d in self._tables_ref]

    def _load_context(self):
        # define here all context
        _tables_ref = []
        wb = self._workbook
        _pivots_ref = []
        for sheet in wb.sheetnames:
            nb_temp = len(_tables_ref)
            sh = wb[sheet]
            _pivots_ref.extend([Cdict(name=pivot.name, sheet=sheet, ref=pivot.location.ref, index=index)
                                for index, pivot in enumerate(sh._pivots)])
            _tables_ref.extend([Cdict(name=k, sheet=sheet, idx=i + nb_temp) for i, k in enumerate(sh.tables.keys())])
        self._tables_ref = _tables_ref
        self._pivots_ref = _pivots_ref

    def table(self, name, get_idx=False, **kwargs):
        if name not in self.tables_name:
            if "default" in kwargs:
                return kwargs["default"]
            raise ValueError("Table %s don't exists" % name)
        for table in self._tables_ref:
            if name == table.name:
                if get_idx:
                    return table.idx, self._workbook[table.sheet].tables[name]
                return self._workbook[table.sheet].tables[name]
        if "default" in kwargs:
            return kwargs["default"]
        raise ValueError("Table %s don't exists" % name)

    def __call__(self, *args, **kwargs):
        if (len(args) and not args[0]) or ("save" in kwargs and not kwargs["save"]):
            self._cm_force_saving = False
        return self

    def __enter__(self):
        return self._workbook

    def __exit__(self, exc_type, exc_val, exc_tb):
        self._load_context()
        self.save(force=self._cm_force_saving)
        self._cm_force_saving = True

    def sheet(self, name, create_if_not_exists=True, sheet_index=None) -> Worksheet:
        if create_if_not_exists and name not in self._workbook.sheetnames:
            self._workbook.create_sheet(name, index=sheet_index)
        return self._workbook[name]

    def save(self, path=None, force=False):
        path = path or self._path
        if not force and not self._got_modification:
            return path
        self._workbook.save(filename=path)
        self._got_modification = False
        return path

    def clear_content(self, ref, sheet=None, format_=True):
        ref, sheet = self._get_ref_sheet_from(ref, sheet)
        _ref = ref.start.address
        if "end" in ref:
            _ref += ":" + ref.end.address
        else:
            sheet[_ref].value = None
            return
        for row in sheet[_ref]:
            for cell in row:
                cell.value = None
                cell.style = 'Normal'

    @staticmethod
    def _parse_ref(ref: str):
        unquoted_ref, quoted_text_dict = replace_quoted_text(ref, quotes="'")
        unquoted_ref = unquoted_ref.split("!")
        if len(unquoted_ref) == 2:
            sheet = unquoted_ref[0]
            ref = unquoted_ref[1]
            for k in quoted_text_dict:
                sheet = sheet.replace(k, quoted_text_dict[k])
        else:
            sheet = None
        ref = ref.replace("$", "")
        res = Cdict()
        if sheet is not None:
            res["sheet"] = sheet
        for ref, part in zip(ref.upper().split(":"), ["start", "end"]):
            _, col, line, _ = re.split(r"^([A-Z]+)?(\d+)?$", ref)
            res[part] = Cdict({"col": col, "line": None if line is None else int(line), "address": ref,
                               "col_id": None if col is None else column_index_from_string(col)})

        return res

    def insert_data_to_table(self, data, table_name, sheet=None):
        if hasattr(data, "shape") and not getattr(data, "shape")[0]:
            return
        elif not hasattr(data, "shape") and not data:
            return

        if sheet is None:
            _, sheet = self._get_ref_sheet_from(table_name)
        elif not isinstance(sheet, Worksheet):
            sheet = self.sheet(sheet)
        table: Table = sheet.tables[table_name]

        data = DatasetFactory(data, columns=table.column_names).dataset
        ref = self._parse_ref(table.ref)
        current_line = ref.end.line + 1
        for row in dataframe_to_rows(data, index=False, header=False):
            current_col = ref.start.col_id
            for value in row:
                sheet.cell(row=current_line, column=current_col).value = value
                current_col += 1
            current_line += 1
        ref = ref.start.col + str(ref.start.line) + ":" + ref.end.col + str(current_line - 1)
        table.ref = ref
        self._got_modification = True

    def copy(self, ref):
        ref = self._parse_ref(ref)
        if ref.get("sheet"):
            sheet = self.sheet(ref.get("sheet"))
        else:
            sheet = self._workbook.active

        self._cache_copy = Cdict()
        if not ref.get("end"):
            self._cache_copy.type = "cell"
            self._cache_copy.ref = {"sheet": sheet.title, "cell": ref.start}
        else:
            self._cache_copy.type = "range"
            ref.start.line = ref.start.line or 1
            ref.end.line = ref.end.line or sheet.max_row

            ref.start.col_id = ref.start.col_id or 1
            ref.end.col_id = ref.end.col_id or sheet.max_column

            ref.start.col = column_index_from_string(ref.start.col_id)
            ref.end.col = column_index_from_string(ref.end.col_id)
            self._cache_copy.ref = {"sheet": sheet.title, "range": ref}

    @staticmethod
    def address(cell: Cell):
        return "'" + cell.parent.title + "'!" + cell.column_letter + str(cell.row)

    @staticmethod
    def clone(cell: Cell, to_cell: Cell, format_=True, data=True, formula_rc=True, consider_format=()):
        if not consider_format:
            consider_format = ("fill", "alignment", "border", "font", "number_format", "protection")
        if data:
            # prise en compte de formula_rc
            to_cell.value = cell.value
        if format_:
            for s in consider_format:
                setattr(to_cell, s, copy(getattr(cell, s)))

    def paste_to(self, ref, format_=True, data=True, ignore_error=True, consider_format=()):
        if not self._cache_copy:
            return
        ref = self._parse_ref(ref)
        cell = ref.start
        if ref.get("sheet"):
            sheet = self.sheet(ref.get("sheet"))
        else:
            sheet = self._workbook.active

        if self._cache_copy.type == "cell":
            cell = sheet.cell(row=cell.line, column=cell.col_id)
            to_cell = self.sheet(self._cache_copy.sheet).cell(self._cache_copy.cell.line,
                                                              self._cache_copy.cell.col_id)
            self.clone(cell, to_cell, format_=format_, data=data, consider_format=consider_format)
        else:
            # range
            start = self._cache_copy.range.start
            end = self._cache_copy.range.start.end

            to_sheet = self.sheet(self._cache_copy.sheet)
            for index_col, col_id in enumerate(range(start.col_id, end.col_id)):
                for index_row, line in enumerate(range(start.line, end.line)):
                    cell = sheet.cell(row=cell.line+index_col, column=cell.col_id+index_row)
                    to_cell = to_sheet.cell(line, col_id)

                    self.clone(cell, to_cell, format_=format_, data=data, consider_format=consider_format)

    def create_table(self, data, ref="A1", sheet=None, table_name=None, header=True,
                     cols_name=None, replace_if_exists=2):
        """
            replace_if_exists: int or bool, bool:: if False and the table_name where exists raise Error; True ==> 1
                                int: 0 -> raise error if table_name where exists
                                     1 -> Replace table content if table where exists
                                     2 -> Adapte table_name if table_name where exists
        """

        if hasattr(data, "shape") and not getattr(data, "shape")[0]:
            return
        elif not hasattr(data, "shape") and not data:
            return
        # Ref calculation
        ref = self._parse_ref(ref)

        if sheet is None:
            sheet = ref.get("sheet")
        # table calculation
        if replace_if_exists == 2:
            if table_name is None:
                table_name = "Tableau"
            new_table_name = table_name
            iter_try = 1
            while new_table_name in self.tables_name:
                new_table_name = table_name + "_" + str(iter_try)
                iter_try += 1
            table_name = new_table_name
        elif table_name is None:
            raise ValueError("table_name arg was omitted")
        replace_if_exists = bool(replace_if_exists)
        table = self.table(table_name, get_idx=True, default=-1)
        if table != -1:
            idx, table = table
            if replace_if_exists:
                self._workbook[self._tables_ref[idx].sheet].tables.pop(table_name)
                last_sheet = self._tables_ref[idx].sheet
                self.clear_content(table.ref, sheet=last_sheet)
                if sheet is None:
                    sheet = last_sheet
            else:
                raise ValueError("Table %s already exists" % table_name)
        if not header and cols_name is None:
            if isinstance(data, dict):
                pass
            elif isinstance(data, list) and isinstance(data[0], list):
                cols_name = ["Colonne_" + str(i) for i in range(len(data[0]))]
        data = DatasetFactory(data, header=header, columns=cols_name).dataset
        if sheet is None:
            sheet = self._workbook.active
        elif not isinstance(sheet, Worksheet):
            sheet = self.sheet(sheet)
        current_line = ref.start.line
        current_col = 1
        for row in dataframe_to_rows(data, index=False, header=True):
            current_col = ref.start.col_id
            for value in row:
                sheet.cell(row=current_line, column=current_col).value = value
                current_col += 1
            current_line += 1

        _ref = ref.start.col + str(ref.start.line) + ":" + get_column_letter(current_col - 1)
        table_ref = _ref + str(current_line - 1)
        table = Table(displayName=table_name, ref=table_ref)
        sheet.add_table(table)
        # context
        self._load_context()
        self._got_modification = True
        return table_name

    def run_macro(self, macro, debug=False):
        path = self.save(force=True)
        import win32com.client as win32
        xlapp = win32.gencache.EnsureDispatch("Excel.Application")  # instantiate excel app
        if debug:
            xlapp.Visible = True
        wb = xlapp.Workbooks.Open(path)
        # xl.Application.Run('(classeur.xlsm!)?Module1.macro1("Jay")')
        xlapp.Application.Run(macro)
        wb.Save()
        xlapp.Application.Quit()
        self.__load()

    def create_macro(self, script: str | list | tuple, module_name=None, run=True, args=None, debug=False):
        global_run = run
        path = self.save(force=True)
        print("File is not saved")
        import win32com.client as win32
        xlapp = win32.gencache.EnsureDispatch("Excel.Application")  # instantiate excel app

        wb = xlapp.Workbooks.Open(path)
        if debug:
            xlapp.Visible = True
        xlmodule = wb.VBProject.VBComponents.Add(1)
        xlmodule.Name = module_name or "kb_Module"
        module_name = xlmodule.Name
        if not isinstance(script, (list, tuple)):
            scripts = [{"run": run, "script": script}]
        else:
            scripts = script
        for script in scripts:
            run = script.get("run", global_run)
            script = script.get("script")
            if not script.strip():
                continue
            xlmodule.CodeModule.AddFromString(script)
            # xl.Application.Run('(classeur.xlsm!)?Module1.macro1("Jay")')
            if run:
                res = re.search(r"^(sub|function)[ \t]+(\w+)[ \t]*(\([\w, \t]*\))$", script.strip().split("\n")[0],
                                flags=re.I | re.S)
                if res:
                    _type, name, _args = res.groups()
                    if _args and args:
                        if isinstance(args, dict):
                            for k in args:
                                pass
                        else:
                            # list, or tuple
                            name += "(" + ",".join([repr(x) for x in args]) + ")"

                    macro = module_name + "." + name
                    xlapp.Application.Run(macro)

        wb.Save()
        xlapp.Application.Quit()
        self.__load()

    def add_sparklines(self, refs, export_name=None, debug=False):
        """
            refs: list or dict.
                if dict: -> {"data_range": str (some sheets range), "destination": Cell or str}
                if list: [dict, ...] where dict like what define top
        """
        import win32com.client as win32
        xlapp = win32.gencache.EnsureDispatch("Excel.Application")  # instantiate excel app
        if debug:
            xlapp.Visible = True
        if isinstance(refs, dict):
            refs = [refs]
        if not len(refs):
            return

        with CTemporaryFile(ext=".xlsx") as temp_file:
            temp_file.close()
            path = self.save(path=temp_file.name, force=True)
            wb = xlapp.Workbooks.Open(path)
            xlmodule = wb.VBProject.VBComponents.Add(1)
            xlmodule.Name = "kb_Module_temp"
            vba_code = "sub create_sparkline_%(index)s()\n" \
                       'sheets("%(sheet)s").Range("%(destination)s").SparklineGroups.Add Type:=xlSparkLine, ' \
                       'SourceData:="%(data_range)s"\n' \
                       'end sub'

            for index, ref in enumerate(refs):

                ref["index"] = index
                if isinstance(ref["destination"], Cell):
                    destination = ExcelFactory.address(ref["destination"])
                else:
                    destination = ref["destination"]
                destination = self._parse_ref(destination)

                ref["sheet"] = destination.get("sheet") or ref.get("sheet") or self._workbook.active.title
                ref["destination"] = destination.start.address
                print(vba_code % ref)
                xlmodule.CodeModule.AddFromString(vba_code % ref)
                xlapp.Application.Run(wb.Name + "!create_sparkline_%s" % (index,))

            wb.Save()
            path = os.path.realpath(export_name or self._path)
            wb.Close(True, path)
            return path

    def refresh_all(self, debug=False):
        import win32com.client as win32
        xlapp = win32.gencache.EnsureDispatch("Excel.Application")  # instantiate excel app
        if debug:
            xlapp.Visible = True
        with CTemporaryFile(ext=".xlsx") as temp_file:
            temp_file.close()
            path = self.save(path=temp_file.name, force=True)
            wb = xlapp.Workbooks.Open(path)
            print(wb)
            wb.RefreshAll()
            print("Saving ...")
            wb.Save()
            wb.Close(True)
            print("Finished")
            # xlapp.Application.Quit()
            self.__load(path)


if __name__ == '__main__':
    excel = ExcelFactory(r"C:\Users\FBYZ6263\Downloads\track_swap.xlsx")
    print(excel.sheet("Report").title)
